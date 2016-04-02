from main.models import *

import openpyxl
print('Opening Workbook...')


wb1 = openpyxl.load_workbook('database.xlsx')
sheet = wb1.get_sheet_by_name('Presiding Officer')
row_count = sheet.get_highest_row()+1

'''
for i in range(2, row_count):
	name = sheet.cell(row=i, column=1).value
	sname = sheet.cell(row=i, column=3).value
	smobile = sheet.cell(row=i, column=4).value
	lac_id = sheet.cell(row=i, column=2).value
	SO = SectorOffice()
	SO.name = name
	lac = LAC.objects.get(unique_id=lac_id)
	SO.lac = lac
	SO.sector_officer = sname
	SO.sector_officer_mobile = smobile
	SO.save()



for i in range(2, row_count):
	unique_id = sheet.cell(row=i, column=2).value
	so_id = sheet.cell(row=i, column=3).value
	so = SectorOffice.objects.get(id=so_id)
	name = sheet.cell(row=i, column=4).value
	lat = sheet.cell(row=i, column=5).value
	lng = sheet.cell(row=i, column=6).value
	PS = PollingStation()
	PS.unique_id = unique_id
	PS.sector_office = so
	PS.name = name
	PS.latitude = lat
	PS.longitude = lng
	PS.save()

for i in range(2, row_count):
	name = sheet.cell(row=i, column=1).value
	mobile = sheet.cell(row=i, column=2).value
	second_mobile = sheet.cell(row=i, column=3).value
	username = sheet.cell(row=i, column=5).value
	ps_id = sheet.cell(row=i, column=4).value
	PS = PollingStation.objects.get(unique_id=ps_id)
	po = PresidingOfficer()
	po.username = username
	po.first_name = name
	po.polling_station = PS
	po.mobile = mobile
	po.second_mobile = second_mobile
	evm = EVM()
	evm.polling_station = PS
	evm.unique_id = username
	po.save()
	evm.save()
'''
print("Saved ;)")